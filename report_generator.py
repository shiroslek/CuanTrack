#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Financial Tracker Bot - Report Generator
v2.4 - Monthly carry-over saldo + charts + notes in PDF
"""

from datetime import datetime, timedelta
import calendar as cal_module

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, Image, PageBreak, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

from database import Database
from calculator import Calculator
from config import EXPORT_DIR, TIMEZONE
from parser import NumberParser
from chart_generator import ChartGenerator


# ── Helper functions ──────────────────────────────────────────

def _get_all_months(user_id, db) -> list:
    """Semua bulan yang ada transaksinya, sorted ascending."""
    db.cursor.execute("""
        SELECT DISTINCT substr(date, 1, 7) as ym
        FROM transactions
        WHERE user_id = ?
        ORDER BY ym ASC
    """, (user_id,))
    result = []
    for row in db.cursor.fetchall():
        ym = row[0]
        result.append((int(ym[:4]), int(ym[5:7])))
    return result


def _month_range(year: int, month: int):
    """(start_date, end_date) untuk satu bulan."""
    start = f"{year:04d}-{month:02d}-01"
    last_day = cal_module.monthrange(year, month)[1]
    end = f"{year:04d}-{month:02d}-{last_day:02d}"
    return start, end


def _month_label(year: int, month: int) -> str:
    bulan_id = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember"
    }
    return f"{bulan_id[month]} {year}"


def _saldo_awal_bulan(user_id, db, year, month) -> int:
    """
    Saldo akhir bulan sebelumnya = semua pemasukan - semua pengeluaran
    s.d. hari terakhir bulan sebelumnya.
    Otomatis jadi saldo awal bulan (year, month).
    """
    first_this = datetime(year, month, 1)
    last_prev = first_this - timedelta(days=1)
    if last_prev.year < 2000:
        return 0
    end_prev = last_prev.strftime("%Y-%m-%d")
    inc = db.get_total_by_type(user_id, 'income', end_date=end_prev)
    exp = db.get_total_by_type(user_id, 'expense', end_date=end_prev)
    return inc - exp


# ── Main class ────────────────────────────────────────────────

class ReportGenerator:
    def __init__(self, db: Database):
        self.db = db
        self.calc = Calculator(db)
        self.parser = NumberParser()
        self.chart_gen = ChartGenerator(db)

    def fmt(self, amount):
        return self.parser.format_rupiah(amount)

    # ── TEXT REPORT ──────────────────────────────────────────

    def generate_text_report(self, user_id, start_date=None, end_date=None) -> str:
        """Laporan teks bulan ini dengan saldo carry-over."""
        now = datetime.now(TIMEZONE)
        if start_date is None:
            start_date = now.replace(day=1).strftime("%Y-%m-%d")
        if end_date is None:
            end_date = now.strftime("%Y-%m-%d")

        sd = datetime.strptime(start_date, "%Y-%m-%d")
        label = _month_label(sd.year, sd.month)

        saldo_awal = _saldo_awal_bulan(user_id, self.db, sd.year, sd.month)
        total_income = self.db.get_total_by_type(user_id, 'income', start_date, end_date)
        total_expense = self.db.get_total_by_type(user_id, 'expense', start_date, end_date)
        saldo_akhir = saldo_awal + total_income - total_expense

        top_cats = self.db.get_spending_by_category(user_id, start_date, end_date)[:5]
        insights = self.calc.generate_insights(user_id)
        recent_income = self.db.get_transactions(user_id, 'income', limit=5,
                                                  start_date=start_date, end_date=end_date)
        recent_expense = self.db.get_transactions(user_id, 'expense', limit=10,
                                                   start_date=start_date, end_date=end_date)
        notes = self.db.get_all_notes(user_id)

        r = f"📊 *LAPORAN KEUANGAN — {label}*\n"
        r += "═══════════════════════\n\n"

        r += "*💰 RINGKASAN*\n"
        r += f"🔄 Saldo Awal Bulan : {self.fmt(saldo_awal)}\n"
        r += f"📥 Pemasukan        : {self.fmt(total_income)}\n"
        r += f"📤 Pengeluaran      : {self.fmt(total_expense)}\n"
        r += "────────────────\n"
        r += f"*💵 Saldo Akhir     : {self.fmt(saldo_akhir)}*\n\n"

        if top_cats:
            r += f"*📊 TOP PENGELUARAN {label.upper()}*\n"
            for i, cat in enumerate(top_cats, 1):
                pct = (cat['total'] / total_expense * 100) if total_expense > 0 else 0
                r += f"{i}. {cat['category']}: {self.fmt(cat['total'])} ({pct:.1f}%)\n"
            r += "\n"

        if insights:
            r += f"*💡 INSIGHT — {label}*\n"
            for i, ins in enumerate(insights, 1):
                r += f"{i}. {ins}\n"
            r += "\n"

        if recent_income:
            r += f"*💰 PEMASUKAN TERAKHIR — {label}*\n"
            for t in recent_income:
                r += f"• {t['date']}: {self.fmt(t['amount'])}\n"
                r += f"  {t['category']} - {t['description'] or '-'}\n"
            r += "\n"

        if recent_expense:
            r += f"*💸 PENGELUARAN TERAKHIR (10) — {label}*\n"
            for t in recent_expense:
                r += f"• {t['date']}: {self.fmt(t['amount'])}\n"
                r += f"  {t['category']} - {t['description'] or '-'}\n"
            r += "\n"

        if notes:
            r += "*📓 NOTES AKTIF*\n"
            for i, note in enumerate(notes, 1):
                r += f"{i}. {note['description']}\n"
            r += "\n"

        r += "═══════════════════════\n"
        r += f"_Generated: {datetime.now(TIMEZONE).strftime('%d %b %Y %H:%M')}_"
        return r

    # ── PDF ──────────────────────────────────────────────────

    def generate_pdf(self, user_id, filename: str = None) -> str:
        if not filename:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'laporan_keuangan_{ts}.pdf'
        filepath = os.path.join(EXPORT_DIR, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                leftMargin=1.8*cm, rightMargin=1.8*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()

        # Custom styles
        S = {
            'title': ParagraphStyle('title', fontSize=18, fontName='Helvetica-Bold',
                                     textColor=colors.HexColor('#2C3E50'),
                                     alignment=TA_CENTER, spaceAfter=6),
            'subtitle': ParagraphStyle('subtitle', fontSize=9,
                                        textColor=colors.grey, alignment=TA_CENTER),
            'h2': ParagraphStyle('h2', fontSize=12, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#2C3E50'),
                                  spaceAfter=6, spaceBefore=10),
            'month_banner': ParagraphStyle('mb', fontSize=14, fontName='Helvetica-Bold',
                                            textColor=colors.white,
                                            alignment=TA_CENTER),
            'normal': styles['Normal'],
        }

        story = []

        # ── Cover ──
        story.append(Paragraph("LAPORAN KEUANGAN PERSONAL", S['title']))
        story.append(Paragraph(
            f"Generated: {datetime.now(TIMEZONE).strftime('%d %B %Y %H:%M')}",
            S['subtitle']
        ))
        story.append(Spacer(1, 10))

        months = _get_all_months(user_id, self.db)
        if not months:
            story.append(Paragraph("Belum ada data transaksi.", S['normal']))
            doc.build(story)
            return filepath

        # ── Ringkasan Keseluruhan ──
        story.append(Paragraph("RINGKASAN KESELURUHAN", S['h2']))
        all_inc = self.db.get_total_by_type(user_id, 'income')
        all_exp = self.db.get_total_by_type(user_id, 'expense')
        all_sal = all_inc - all_exp

        overall_rows = [
            ['Bulan', 'Saldo Awal', 'Pemasukan', 'Pengeluaran', 'Saldo Akhir'],
        ]
        for y, m in months:
            s_date, e_date = _month_range(y, m)
            lbl = _month_label(y, m)
            s_awal = _saldo_awal_bulan(user_id, self.db, y, m)
            inc = self.db.get_total_by_type(user_id, 'income', s_date, e_date)
            exp = self.db.get_total_by_type(user_id, 'expense', s_date, e_date)
            sal = s_awal + inc - exp
            overall_rows.append([lbl, self.fmt(s_awal), self.fmt(inc),
                                  self.fmt(exp), self.fmt(sal)])
        overall_rows.append(['SALDO SAAT INI', '', self.fmt(all_inc),
                              self.fmt(all_exp), self.fmt(all_sal)])

        ot = Table(overall_rows, colWidths=[3.8*cm, 3.2*cm, 3.2*cm, 3.2*cm, 3.2*cm])
        ot.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8.5),
            ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.HexColor('#EBF5FB'), colors.white]),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#D6EAF8')),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ]))
        story.append(ot)
        story.append(Spacer(1, 10))

        # ── Notes global (tampil sekali di awal) ──
        notes = self.db.get_all_notes(user_id)
        if notes:
            story.append(Paragraph("CATATAN (NOTES)", S['h2']))
            notes_data = [['No', 'Catatan', 'Waktu']]
            for i, note in enumerate(notes, 1):
                notes_data.append([str(i), note['description'],
                                    str(note['created_at'])[:16]])
            nt = Table(notes_data, colWidths=[1*cm, 13*cm, 3.6*cm])
            nt.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F39C12')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8.5),
                ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('LEFTPADDING', (0,0), (-1,-1), 6),
                ('ROWBACKGROUNDS', (0,1), (-1,-1),
                 [colors.lightgoldenrodyellow, colors.white]),
            ]))
            story.append(nt)

        story.append(PageBreak())

        # ── Per bulan ──
        for year, month in months:
            start_date, end_date = _month_range(year, month)
            label = _month_label(year, month)

            saldo_awal = _saldo_awal_bulan(user_id, self.db, year, month)
            inc_txs = self.db.get_transactions(user_id, 'income',
                                               start_date=start_date, end_date=end_date)
            exp_txs = self.db.get_transactions(user_id, 'expense',
                                               start_date=start_date, end_date=end_date)
            total_inc = self.db.get_total_by_type(user_id, 'income', start_date, end_date)
            total_exp = self.db.get_total_by_type(user_id, 'expense', start_date, end_date)
            saldo_akhir = saldo_awal + total_inc - total_exp

            # Month banner
            banner = Table(
                [[Paragraph(f"📅  {label.upper()}", S['month_banner'])]],
                colWidths=[16.2*cm]
            )
            banner.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#2C3E50')),
                ('TOPPADDING', (0,0), (-1,-1), 10),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(banner)
            story.append(Spacer(1, 8))

            # Summary bulan — saldo awal → pemasukan → pengeluaran → saldo akhir
            sum_data = [
                ['Keterangan', 'Jumlah'],
                ['🔄 Saldo Awal Bulan', self.fmt(saldo_awal)],
                ['📥 Pemasukan', self.fmt(total_inc)],
                ['📤 Pengeluaran', self.fmt(total_exp)],
                ['💵 Saldo Akhir', self.fmt(saldo_akhir)],
            ]
            st = Table(sum_data, colWidths=[8*cm, 8*cm])
            st.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#5D6D7E')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 10),
                ('GRID', (0,0), (-1,-1), 0.4, colors.lightgrey),
                ('TOPPADDING', (0,0), (-1,-1), 7),
                ('BOTTOMPADDING', (0,0), (-1,-1), 7),
                ('LEFTPADDING', (0,0), (-1,-1), 10),
                ('ROWBACKGROUNDS', (0,1), (-1,-2),
                 [colors.HexColor('#EBF5FB'), colors.white,
                  colors.HexColor('#FDEDEC'), colors.white]),
                ('BACKGROUND', (0,4), (-1,4), colors.HexColor('#D6EAF8')),
                ('FONTNAME', (0,4), (-1,4), 'Helvetica-Bold'),
            ]))
            story.append(st)
            story.append(Spacer(1, 10))

            # Tabel Pemasukan
            if inc_txs:
                story.append(Paragraph(f"Pemasukan — {label}", S['h2']))
                inc_data = [['Tanggal', 'Kategori', 'Keterangan', 'Jumlah']]
                for t in inc_txs:
                    inc_data.append([t['date'], t['category'],
                                     t['description'] or '-', self.fmt(t['amount'])])
                inc_data.append(['', '', 'TOTAL PEMASUKAN', self.fmt(total_inc)])

                it = Table(inc_data, colWidths=[2.8*cm, 3.8*cm, 6.2*cm, 3.4*cm])
                it.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#27AE60')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,-1), 8.5),
                    ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
                    ('TOPPADDING', (0,0), (-1,-1), 5),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                    ('LEFTPADDING', (0,0), (-1,-1), 6),
                    ('ROWBACKGROUNDS', (0,1), (-1,-2),
                     [colors.HexColor('#EAFAF1'), colors.white]),
                    ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#D5F4E6')),
                    ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
                ]))
                story.append(it)
                story.append(Spacer(1, 8))

            # Tabel Pengeluaran
            if exp_txs:
                story.append(Paragraph(f"Pengeluaran — {label}", S['h2']))
                exp_data = [['Tanggal', 'Kategori', 'Keterangan', 'Jumlah']]
                for t in exp_txs:
                    exp_data.append([t['date'], t['category'],
                                     t['description'] or '-', self.fmt(t['amount'])])
                exp_data.append(['', '', 'TOTAL PENGELUARAN', self.fmt(total_exp)])

                et = Table(exp_data, colWidths=[2.8*cm, 3.8*cm, 6.2*cm, 3.4*cm])
                et.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E74C3C')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,-1), 8.5),
                    ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
                    ('TOPPADDING', (0,0), (-1,-1), 5),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                    ('LEFTPADDING', (0,0), (-1,-1), 6),
                    ('ROWBACKGROUNDS', (0,1), (-1,-2),
                     [colors.HexColor('#FDEDEC'), colors.white]),
                    ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#FADBD8')),
                    ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
                ]))
                story.append(et)
                story.append(Spacer(1, 8))

            # Grafik per bulan
            try:
                try:
                    pie_file = self.chart_gen.generate_expense_pie_chart(
                        filename=f"pie_exp_{year}{month:02d}.png"
                    )
                except TypeError:
                    pie_file = self.chart_gen.generate_expense_pie_chart()
                if pie_file and os.path.exists(pie_file):
                    story.append(Paragraph(
                        "Distribusi Pengeluaran per Kategori", S['h2']))
                    story.append(Image(pie_file, width=14*cm, height=10*cm))
                    story.append(Spacer(1, 8))
            except Exception as e:
                print(f"Pie chart error: {e}")

            try:
                try:
                    trend_file = self.chart_gen.generate_trend_chart(
                        filename=f"trend_{year}{month:02d}.png"
                    )
                except TypeError:
                    trend_file = self.chart_gen.generate_trend_chart()
                if trend_file and os.path.exists(trend_file):
                    story.append(Paragraph(
                        "Trend Pengeluaran Harian (30 Hari Terakhir)", S['h2']))
                    story.append(Image(trend_file, width=14*cm, height=8*cm))
            except Exception as e:
                print(f"Trend chart error: {e}")

            story.append(PageBreak())

        doc.build(story)
        return filepath

    # ── EXCEL ─────────────────────────────────────────────────

    def generate_excel(self, user_id, filename: str = None) -> str:
        if not filename:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'laporan_keuangan_{ts}.xlsx'
        filepath = os.path.join(EXPORT_DIR, filename)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        months = _get_all_months(user_id, self.db)
        self._excel_ringkasan(wb, user_id, months)
        for y, m in months:
            s_date, e_date = _month_range(y, m)
            self._excel_month_sheet(wb, user_id, y, m, s_date, e_date)
        self._excel_notes(wb, user_id)

        wb.save(filepath)
        return filepath

    def _excel_ringkasan(self, wb, user_id, months):
        ws = wb.create_sheet('Ringkasan')
        ws['A1'] = 'RINGKASAN KESELURUHAN'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')

        hdrs = ['Bulan', 'Saldo Awal', 'Pemasukan', 'Pengeluaran', 'Saldo Akhir']
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(3, c, h)
            cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')

        all_inc = self.db.get_total_by_type(user_id, 'income')
        all_exp = self.db.get_total_by_type(user_id, 'expense')

        row = 4
        for y, m in months:
            s, e = _month_range(y, m)
            sa = _saldo_awal_bulan(user_id, self.db, y, m)
            inc = self.db.get_total_by_type(user_id, 'income', s, e)
            exp = self.db.get_total_by_type(user_id, 'expense', s, e)
            sal = sa + inc - exp
            ws.cell(row, 1, _month_label(y, m))
            for c, v in [(2, sa), (3, inc), (4, exp), (5, sal)]:
                ws.cell(row, c, v).number_format = '#,##0'
            row += 1

        ws.cell(row, 1, 'TOTAL').font = Font(bold=True)
        for c, v in [(3, all_inc), (4, all_exp), (5, all_inc - all_exp)]:
            cell = ws.cell(row, c, v)
            cell.number_format = '#,##0'
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)

        for col, w in zip('ABCDE', [20, 18, 18, 18, 18]):
            ws.column_dimensions[col].width = w

    def _excel_month_sheet(self, wb, user_id, year, month, s_date, e_date):
        label = _month_label(year, month)
        ws = wb.create_sheet(label[:31])

        saldo_awal = _saldo_awal_bulan(user_id, self.db, year, month)
        inc_txs = self.db.get_transactions(user_id, 'income',
                                           start_date=s_date, end_date=e_date)
        exp_txs = self.db.get_transactions(user_id, 'expense',
                                           start_date=s_date, end_date=e_date)
        total_inc = self.db.get_total_by_type(user_id, 'income', s_date, e_date)
        total_exp = self.db.get_total_by_type(user_id, 'expense', s_date, e_date)
        saldo_akhir = saldo_awal + total_inc - total_exp

        r = 1
        ws.cell(r, 1, f'LAPORAN {label.upper()}').font = Font(size=13, bold=True)
        ws.merge_cells(f'A{r}:D{r}')
        r += 2

        # Summary
        for lbl, val, color in [
            ('🔄 Saldo Awal Bulan', saldo_awal, None),
            ('📥 Pemasukan', total_inc, '27AE60'),
            ('📤 Pengeluaran', total_exp, 'E74C3C'),
            ('💵 Saldo Akhir', saldo_akhir, '3498DB'),
        ]:
            ws.cell(r, 1, lbl).font = Font(bold=True)
            c = ws.cell(r, 2, val)
            c.number_format = '#,##0'
            if color:
                c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                c.font = Font(color='FFFFFF', bold=True)
            r += 1
        r += 1

        # Pemasukan
        if inc_txs:
            ws.cell(r, 1, 'PEMASUKAN').font = Font(size=11, bold=True)
            ws.merge_cells(f'A{r}:D{r}')
            r += 1
            for c_idx, h in enumerate(['Tanggal', 'Kategori', 'Keterangan', 'Jumlah'], 1):
                cell = ws.cell(r, c_idx, h)
                cell.fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            r += 1
            for t in inc_txs:
                ws.cell(r, 1, t['date'])
                ws.cell(r, 2, t['category'])
                ws.cell(r, 3, t['description'] or '-')
                ws.cell(r, 4, t['amount']).number_format = '#,##0'
                r += 1
            ws.cell(r, 3, 'TOTAL').font = Font(bold=True)
            c = ws.cell(r, 4, total_inc)
            c.number_format = '#,##0'
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')
            r += 2

        # Pengeluaran
        if exp_txs:
            ws.cell(r, 1, 'PENGELUARAN').font = Font(size=11, bold=True)
            ws.merge_cells(f'A{r}:D{r}')
            r += 1
            for c_idx, h in enumerate(['Tanggal', 'Kategori', 'Keterangan', 'Jumlah'], 1):
                cell = ws.cell(r, c_idx, h)
                cell.fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            r += 1
            for t in exp_txs:
                ws.cell(r, 1, t['date'])
                ws.cell(r, 2, t['category'])
                ws.cell(r, 3, t['description'] or '-')
                ws.cell(r, 4, t['amount']).number_format = '#,##0'
                r += 1
            ws.cell(r, 3, 'TOTAL').font = Font(bold=True)
            c = ws.cell(r, 4, total_exp)
            c.number_format = '#,##0'
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')

        for col, w in zip('ABCD', [12, 20, 35, 15]):
            ws.column_dimensions[col].width = w

    def _excel_notes(self, wb, user_id):
        ws = wb.create_sheet('Catatan')
        ws['A1'] = 'DAFTAR CATATAN'
        ws['A1'].font = Font(size=13, bold=True)
        ws.merge_cells('A1:C1')

        for c_idx, h in enumerate(['No', 'Catatan', 'Waktu'], 1):
            cell = ws.cell(3, c_idx, h)
            cell.fill = PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)

        notes = self.db.get_all_notes(user_id)
        for row, note in enumerate(notes, 4):
            ws.cell(row, 1, row - 3)
            ws.cell(row, 2, note['description'])
            ws.cell(row, 3, str(note['created_at'])[:16])

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 18
